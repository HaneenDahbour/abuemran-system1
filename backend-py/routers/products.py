import json
from decimal import Decimal
from datetime import date, datetime
from typing import Optional, Dict, Any
from uuid import UUID
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Query
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from pydantic import BaseModel, Field

from config.db import get_pool
from middleware.auth import get_current_user
from middleware.roles import require_role


router = APIRouter()


def safe_uuid(val):
    try:
        return UUID(str(val))
    except (ValueError, AttributeError, TypeError):
        return None


def to_json_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def row_to_dict(row):
    return {key: to_json_value(row[key]) for key in row.keys()}


def clean_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None

    value = str(value).strip()
    return value or None


def require_warehouse_editor(user):
    require_role(user, "admin", "accountant", "employee")


async def insert_audit(conn, user, action: str, detail: str):
    try:
        await conn.execute(
            """
            INSERT INTO audit_log (user_id, user_name, action, detail)
            VALUES ($1, $2, $3, $4)
            """,
            user.get("id"),
            user.get("full_name") or user.get("username") or "مستخدم",
            action,
            detail,
        )
    except Exception:
        pass


async def ensure_category_exists(conn, category_id: Optional[int]):
    if category_id is None:
        return

    exists = await conn.fetchval(
        "SELECT EXISTS(SELECT 1 FROM warehouse_categories WHERE id=$1)",
        category_id,
    )

    if not exists:
        raise HTTPException(
            status_code=400, detail="الفئة المختارة غير موجودة")


async def ensure_sku_unique(conn, sku: Optional[str], exclude_product_id: Optional[UUID] = None):
    sku = clean_text(sku)

    if not sku:
        return

    if exclude_product_id:
        row = await conn.fetchrow(
            """
            SELECT id, name
            FROM products
            WHERE LOWER(sku) = LOWER($1)
              AND id <> $2
            LIMIT 1
            """,
            sku,
            exclude_product_id,
        )
    else:
        row = await conn.fetchrow(
            """
            SELECT id, name
            FROM products
            WHERE LOWER(sku) = LOWER($1)
            LIMIT 1
            """,
            sku,
        )

    if row:
        raise HTTPException(
            status_code=400,
            detail=f"الكود مستخدم مسبقاً في الصنف: {row['name']}",
        )


class ProductRequest(BaseModel):
    name: str
    sku: Optional[str] = None
    category_id: Optional[int] = None
    unit: Optional[str] = "قطعة"
    min_stock: Optional[float] = 0
    cost_price: Optional[float] = 0       # سعر التكلفة / الشراء
    base_price: Optional[float] = None    # legacy alias — ignored if cost_price provided
    properties: Optional[Dict[str, Any]] = Field(default_factory=dict)
    opening_quantity: Optional[float] = 0
    final_stock: Optional[float] = None


class AdjustRequest(BaseModel):
    quantity: float
    notes: Optional[str] = None


def to_float_or_zero(value):
    if value is None or value == "":
        return 0.0

    try:
        if isinstance(value, str):
            value = value.strip()
            value = value.replace(",", "")
            value = value.replace("٫", ".")
            value = value.replace("،", "")
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def excel_text(value):
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def excel_json_value(value):
    """
    Make Excel values safe for json.dumps.
    """
    if value is None:
        return None

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (date, datetime)):
        return value.isoformat()

    if isinstance(value, str):
        text = value.strip()
        return text or None

    return value


def clean_excel_row(headers, row):
    """
    Build row dict while ignoring empty header columns.
    """
    item = {}

    for header, value in zip(headers, row):
        header = excel_text(header)

        if not header:
            continue

        item[header] = excel_json_value(value)

    return item


def build_import_product_name(item: Dict[str, Any]) -> str:
    """
    Build a readable unique product name from Excel.
    """
    base = excel_text(item.get("product_name")) or ""
    extra_parts = []

    for key in ["brand_or_type", "model", "color", "size"]:
        value = excel_text(item.get(key))

        if value and value not in base:
            extra_parts.append(value)

    if extra_parts:
        return f"{base} - {' - '.join(extra_parts)}"

    return base


async def get_or_create_warehouse_category(conn, category_name: str) -> int:
    category_name = clean_text(category_name)

    if not category_name:
        raise HTTPException(
            status_code=400, detail="اسم الفئة في ملف الإكسل مفقود")

    existing = await conn.fetchrow(
        """
        SELECT id
        FROM warehouse_categories
        WHERE LOWER(name) = LOWER($1)
        LIMIT 1
        """,
        category_name,
    )

    if existing:
        return existing["id"]

    created = await conn.fetchrow(
        """
        INSERT INTO warehouse_categories (name)
        VALUES ($1)
        RETURNING id
        """,
        category_name,
    )

    return created["id"]


def parse_product_properties(value):
    if value is None:
        return {}

    if isinstance(value, dict):
        return value

    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return {}

    return {}


def normalize_arabic(value):
    text = str(value or "").strip()
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    text = text.replace("ى", "ي").replace("ة", "ه")
    return text


def detect_export_category(category_name):
    text = normalize_arabic(category_name)

    if "مكيف" in text:
        return "مكيفات"

    if "كروكس" in text:
        return "كروكس"

    if "ملابس" in text:
        return "ملابس"

    if "حذ" in text or "احذيه" in text or "احذية" in text:
        return "الاحذية"

    if "ادوات" in text or "منزليه" in text or "منزلية" in text:
        return "الادوات المنزلية"

    return "اخرى"


def clean_excel_number(value):
    number = to_float_or_zero(value)

    if abs(number - round(number)) < 0.000001:
        return int(round(number))

    return round(number, 3)


def get_prop(props, *keys):
    for key in keys:
        value = props.get(key)
        if value is not None and value != "":
            return value
    return None


def calc_original_total(props, current_stock):
    total = to_float_or_zero(props.get("original_total_qty"))

    if total > 0:
        return total

    containers = to_float_or_zero(props.get("containers_count"))
    pack_size = to_float_or_zero(props.get("pack_size"))
    loose = to_float_or_zero(props.get("loose_count"))

    if containers > 0 and pack_size > 0:
        return containers * pack_size + loose

    return to_float_or_zero(current_stock)


def calc_sold_qty(props, current_stock, out_qty):
    original_total = calc_original_total(props, current_stock)
    current = to_float_or_zero(current_stock)

    if original_total > 0 and original_total >= current:
        return original_total - current

    sold_before = to_float_or_zero(props.get("sold_qty_before_system"))
    sold_after = to_float_or_zero(out_qty)

    return sold_before + sold_after


def style_export_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.###"

    for col in range(1, ws.max_column + 1):
        max_length = 12

        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_length = max(max_length, len(str(value)) + 3)

        ws.column_dimensions[get_column_letter(
            col)].width = min(max_length, 28)

    ws.auto_filter.ref = ws.dimensions

@router.get("")
@router.get("/")
async def get_products(user=Depends(get_current_user)):
    pool = await get_pool()

    try:
        rows = await pool.fetch(
            """
            SELECT p.*, wc.name AS category_name, wc.icon AS category_icon
            FROM products p
            LEFT JOIN warehouse_categories wc ON wc.id = p.category_id
            ORDER BY wc.name NULLS LAST, p.name
            """
        )

        return [row_to_dict(r) for r in rows]

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("")
@router.post("/")
async def create_product(data: ProductRequest, user=Depends(get_current_user)):
    require_warehouse_editor(user)

    name = clean_text(data.name)
    sku = clean_text(data.sku)
    unit = clean_text(data.unit) or "قطعة"
    min_stock = float(data.min_stock or 0)
    opening_quantity = float(data.opening_quantity or 0)
    base_price = round(float(data.cost_price if data.cost_price is not None else (data.base_price or 0)), 3)
    category_id = int(data.category_id) if data.category_id else None

    if not name:
        raise HTTPException(status_code=400, detail="اسم الصنف مطلوب")

    if min_stock < 0:
        raise HTTPException(
            status_code=400, detail="الحد الأدنى لا يمكن أن يكون سالباً")

    if opening_quantity < 0:
        raise HTTPException(
            status_code=400, detail="الكمية الحالية لا يمكن أن تكون سالبة")

    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                await ensure_category_exists(conn, category_id)
                await ensure_sku_unique(conn, sku)

                category_name = None

                if category_id:
                    category_name = await conn.fetchval(
                        "SELECT name FROM warehouse_categories WHERE id=$1",
                        category_id,
                    )

                duplicate = await conn.fetchrow(
                    """
                    SELECT id, name
                    FROM products
                    WHERE LOWER(name) = LOWER($1)
                      AND COALESCE(category_id, 0) = COALESCE($2::int, 0)
                    LIMIT 1
                    """,
                    name,
                    category_id,
                )

                if duplicate:
                    raise HTTPException(
                        status_code=409,
                        detail=f"الصنف موجود مسبقاً في هذه الفئة: {duplicate['name']}",
                    )

                props_json = json.dumps(
                    data.properties or {}, ensure_ascii=False)

                row = await conn.fetchrow(
                    """
                    INSERT INTO products
                    (name, sku, category_id, category, unit, min_stock, current_stock, properties, cost_price, base_price)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8::jsonb,$9,$9)
                    RETURNING *
                    """,
                    name, sku, category_id, category_name, unit, min_stock,
                    opening_quantity, props_json, base_price,
                )

                if opening_quantity > 0:
                    await conn.execute(
                        """
                        INSERT INTO stock_movements
                          (product_id, type, quantity, source_type, notes, created_by)
                        VALUES
                          ($1, $2, $3, 'opening', $4, $5)
                        """,
                        row["id"],
                        "in",
                        opening_quantity,
                        f"كمية افتتاحية: {opening_quantity} {unit}",
                        user.get("id"),
                    )

                await insert_audit(
                    conn,
                    user,
                    "أضاف صنف مستودع",
                    f"{name} — الكمية الافتتاحية: {opening_quantity} {unit}",
                )

                return row_to_dict(row)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import-excel")
async def import_products_excel(
    file: UploadFile = File(...),
    update_existing: bool = False,
    user=Depends(get_current_user),
):
    require_warehouse_editor(user)

    if not file.filename:
        raise HTTPException(status_code=400, detail="اسم الملف غير موجود")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="الرجاء رفع ملف Excel بصيغة xlsx فقط")

    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="ملف الإكسل فارغ")

    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ملف الإكسل غير صالح")

    if "import_ready_products" not in workbook.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="ارفع الملف الجاهز للاستيراد الذي يحتوي على صفحة import_ready_products",
        )

    sheet = workbook["import_ready_products"]

    if sheet.max_row < 2:
        raise HTTPException(
            status_code=400, detail="ملف الإكسل لا يحتوي على أصناف للاستيراد")

    headers = [excel_text(cell.value) for cell in sheet[1]]

    required_columns = ["category", "product_name", "current_stock_qty"]

    for col in required_columns:
        if col not in headers:
            raise HTTPException(
                status_code=400,
                detail=f"العمود المطلوب غير موجود في الإكسل: {col}",
            )

    inserted = 0
    updated = 0
    skipped = 0
    failed = []

    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    item = clean_excel_row(headers, row)

                    category_name = excel_text(item.get("category"))
                    product_name = build_import_product_name(item)
                    current_stock = to_float_or_zero(
                        item.get("current_stock_qty"))

                    if not category_name or not product_name:
                        skipped += 1
                        failed.append({
                            "row": row_index,
                            "reason": "الفئة أو اسم الصنف مفقود",
                        })
                        continue

                    if current_stock < 0:
                        skipped += 1
                        failed.append({
                            "row": row_index,
                            "product": product_name,
                            "reason": "كمية المخزون سالبة",
                        })
                        continue

                    properties = {
                        "import_source": "excel",
                        "category_from_excel": category_name,
                        "base_product_name": excel_text(item.get("product_name")),
                        "brand_or_type": excel_text(item.get("brand_or_type")),
                        "model": excel_text(item.get("model")),
                        "color": excel_text(item.get("color")),
                        "size": excel_text(item.get("size")),
                        "pack_size": excel_json_value(item.get("pack_size")),
                        "containers_count": excel_json_value(item.get("containers_count")),
                        "loose_count": excel_json_value(item.get("loose_count")),
                        "original_total_qty": excel_json_value(item.get("original_total_qty")),
                        "sold_qty_before_system": excel_json_value(item.get("sold_qty")),
                        "current_stock_qty_imported": current_stock,
                        "unit_price_from_excel": excel_json_value(item.get("unit_price")),
                        "total_price_from_excel": excel_json_value(item.get("total_price")),
                        "import_note": excel_text(item.get("import_note")),
                    }

                    props_json = json.dumps(properties, ensure_ascii=False)

                    async with conn.transaction():
                        category_id = await get_or_create_warehouse_category(conn, category_name)

                        existing = await conn.fetchrow(
                            """
                            SELECT id, name, current_stock
                            FROM products
                            WHERE LOWER(name) = LOWER($1)
                              AND COALESCE(category_id, 0) = COALESCE($2::int, 0)
                            LIMIT 1
                            """,
                            product_name,
                            category_id,
                        )

                        if existing:
                            if update_existing:
                                old_stock = float(
                                    existing["current_stock"] or 0)
                                diff = current_stock - old_stock

                                await conn.execute(
                                    """
                                    UPDATE products
                                    SET current_stock=$1,
                                        category=$2,
                                        properties=$3::jsonb
                                    WHERE id=$4
                                    """,
                                    current_stock,
                                    category_name,
                                    props_json,
                                    existing["id"],
                                )

                                if diff != 0:
                                    await conn.execute(
                                        """
                                        INSERT INTO stock_movements
                                          (product_id, type, quantity, source_type, notes, created_by)
                                        VALUES
                                          ($1, $2, $3, 'excel_import_update', $4, $5)
                                        """,
                                        existing["id"],
                                        "in" if diff > 0 else "out",
                                        abs(diff),
                                        f"تحديث من Excel: من {old_stock} إلى {current_stock}",
                                        user.get("id"),
                                    )

                                updated += 1
                            else:
                                skipped += 1

                            continue

                        created = await conn.fetchrow(
                            """
                            INSERT INTO products
                              (name, sku, category_id, category, unit, min_stock, current_stock, properties)
                            VALUES
                              ($1, NULL, $2, $3, $4, $5, $6, $7::jsonb)
                            RETURNING id
                            """,
                            product_name,
                            category_id,
                            category_name,
                            "قطعة",
                            0,
                            current_stock,
                            props_json,
                        )

                        if current_stock > 0:
                            await conn.execute(
                                """
                                INSERT INTO stock_movements
                                  (product_id, type, quantity, source_type, notes, created_by)
                                VALUES
                                  ($1, 'in', $2, 'excel_import', $3, $4)
                                """,
                                created["id"],
                                current_stock,
                                f"استيراد Excel — مخزون افتتاحي: {current_stock}",
                                user.get("id"),
                            )

                        inserted += 1

                except Exception as row_error:
                    skipped += 1
                    failed.append({
                        "row": row_index,
                        "reason": str(row_error),
                    })
                    continue

            # ✅ مهم: سجل التدقيق داخل نفس block تبع conn
            try:
                await insert_audit(
                    conn,
                    user,
                    "استيراد أصناف من Excel",
                    f"تمت إضافة {inserted} صنف، تحديث {updated} صنف، تخطي {skipped} صنف",
                )
            except Exception as audit_error:
                print("Audit log failed after Excel import:", audit_error)

        return {
            "success": True,
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "failed": failed[:50],
            "message": f"تم الاستيراد: إضافة {inserted}، تحديث {updated}، تخطي {skipped}",
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-excel")
async def export_products_excel(user=Depends(get_current_user)):
    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                """
                SELECT
                    p.id,
                    p.name,
                    p.sku,
                    p.category,
                    p.current_stock,
                    p.properties,
                    wc.name AS category_name,
                    COALESCE(out_movements.out_qty, 0) AS out_qty
                FROM products p
                LEFT JOIN warehouse_categories wc ON wc.id = p.category_id
                LEFT JOIN (
                    SELECT
                        product_id,
                        SUM(quantity) AS out_qty
                    FROM stock_movements
                    WHERE type = 'out'
                    GROUP BY product_id
                ) out_movements ON out_movements.product_id = p.id
                ORDER BY wc.name NULLS LAST, p.name ASC
                """
            )

            await insert_audit(
                conn,
                user,
                "تصدير أصناف المستودع Excel",
                f"تم تصدير {len(rows)} صنف من المستودع",
            )

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        sheet_headers = {
            "مكيفات": [
                "النوع",
                "الصنف",
                "العدد الموجود",
                "المباع",
                "الباقي",
                "السعر",
            ],
            "كروكس": [
                "الصنف",
                "الموديل",
                "الكمية",
                "عدد الشوالات",
                "العدد الافرادي",
                "المجموع كامل",
                "المباع",
                "الباقي",
                "السعر الافرادي",
                "السعر الاجمالي",
            ],
            "ملابس": [
                "الصنف",
                "الموديل",
                "الكمية",
                "عدد الشوالات",
                "العدد الافرادي",
                "العدد الكلي",
                "المباع",
                "الباقي",
            ],
            "الاحذية": [
                "الصنف",
                "رقم الموديل",
                "اللون",
                "المقاس",
                "التعبئة",
                "عدد الكراتين",
                "العدد الافرادي",
                "المباع",
                "الباقي",
            ],
            "الادوات المنزلية": [
                "الصنف",
                "عدد الكراتين",
                "تعبئة الكراتين",
                "العدد الافرادي",
                "المباع",
                "الباقي",
            ],
            "اخرى": [
                "الفئة",
                "الصنف",
                "الكود",
                "الكمية الحالية",
                "الوحدة",
            ],
        }

        sheets = {}

        for sheet_name, headers in sheet_headers.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            sheets[sheet_name] = ws

        for row in rows:
            props = parse_product_properties(row["properties"])

            category_name = (
                row["category_name"]
                or row["category"]
                or props.get("category_from_excel")
                or "اخرى"
            )

            export_category = detect_export_category(category_name)

            base_product_name = (
                props.get("base_product_name")
                or props.get("product_name")
                or row["name"]
            )

            brand_or_type = get_prop(props, "brand_or_type", "type")
            model = get_prop(props, "model", "رقم الموديل")
            color = get_prop(props, "color", "لون")
            size = get_prop(props, "size", "مقاس")

            pack_size = clean_excel_number(props.get("pack_size"))
            containers_count = clean_excel_number(
                props.get("containers_count"))
            loose_count = clean_excel_number(props.get("loose_count"))

            original_total = clean_excel_number(
                calc_original_total(props, row["current_stock"])
            )

            sold_qty = clean_excel_number(
                calc_sold_qty(props, row["current_stock"], row["out_qty"])
            )

            remaining_qty = clean_excel_number(row["current_stock"])

            unit_price = props.get("unit_price_from_excel")
            total_price = props.get("total_price_from_excel")

            if export_category == "مكيفات":
                sheets["مكيفات"].append([
                    brand_or_type,
                    base_product_name,
                    original_total,
                    sold_qty,
                    remaining_qty,
                    unit_price,
                ])

            elif export_category == "كروكس":
                sheets["كروكس"].append([
                    base_product_name,
                    model,
                    pack_size,
                    containers_count,
                    loose_count,
                    original_total,
                    sold_qty,
                    remaining_qty,
                    unit_price,
                    total_price,
                ])

            elif export_category == "ملابس":
                sheets["ملابس"].append([
                    base_product_name,
                    model,
                    pack_size,
                    containers_count,
                    loose_count,
                    original_total,
                    sold_qty,
                    remaining_qty,
                ])

            elif export_category == "الاحذية":
                sheets["الاحذية"].append([
                    base_product_name,
                    model,
                    color,
                    size,
                    pack_size,
                    containers_count,
                    original_total,
                    sold_qty,
                    remaining_qty,
                ])

            elif export_category == "الادوات المنزلية":
                sheets["الادوات المنزلية"].append([
                    base_product_name,
                    containers_count,
                    pack_size,
                    original_total,
                    sold_qty,
                    remaining_qty,
                ])

            else:
                sheets["اخرى"].append([
                    category_name,
                    row["name"],
                    row["sku"],
                    remaining_qty,
                    "قطعة",
                ])

        for ws in wb.worksheets:
            style_export_sheet(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"تصدير-المستودع-{datetime.now().strftime('%Y-%m-%d-%H-%M')}.xlsx"
        encoded_filename = quote(filename)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/profit-analysis")
async def get_profit_analysis(user=Depends(get_current_user)):
    require_role(user, "admin", "accountant")
    pool = await get_pool()
    try:
        rows = await pool.fetch("""
            SELECT
                p.id, p.name, p.sku, p.unit,
                COALESCE(p.cost_price, p.base_price, 0) AS cost_price,
                p.current_stock,
                wc.name                                  AS category_name,
                COALESCE(s.total_sold_qty, 0)            AS total_sold_qty,
                COALESCE(s.total_revenue, 0)             AS total_revenue,
                CASE WHEN COALESCE(s.total_sold_qty,0) > 0
                     THEN ROUND((s.total_revenue / s.total_sold_qty)::numeric, 3)
                     ELSE 0 END                          AS avg_selling_price,
                CASE WHEN COALESCE(p.cost_price, p.base_price, 0) > 0
                          AND COALESCE(s.total_sold_qty,0) > 0
                     THEN ROUND((s.total_revenue/s.total_sold_qty
                          - COALESCE(p.cost_price, p.base_price, 0))::numeric, 3)
                     ELSE 0 END                          AS profit_per_unit,
                CASE WHEN COALESCE(p.cost_price, p.base_price, 0) > 0
                          AND COALESCE(s.total_sold_qty,0) > 0
                     THEN ROUND(
                       ((s.total_revenue/s.total_sold_qty
                          - COALESCE(p.cost_price, p.base_price, 0))
                        / COALESCE(p.cost_price, p.base_price, 1) * 100)::numeric, 2)
                     ELSE 0 END                          AS profit_margin_pct,
                CASE WHEN COALESCE(p.cost_price, p.base_price, 0) > 0
                          AND COALESCE(s.total_sold_qty,0) > 0
                     THEN COALESCE(s.total_sold_qty,0)
                          * (s.total_revenue/s.total_sold_qty
                             - COALESCE(p.cost_price, p.base_price, 0))
                     ELSE 0 END                          AS total_profit
            FROM products p
            LEFT JOIN warehouse_categories wc ON wc.id = p.category_id
            LEFT JOIN (
                SELECT ii.product_id,
                       SUM(ii.quantity) AS total_sold_qty,
                       SUM(COALESCE(ii.line_total, ii.quantity * ii.unit_price, 0)) AS total_revenue
                FROM invoice_items ii
                JOIN invoices i ON i.id = ii.invoice_id
                WHERE COALESCE(NULLIF(i.status,''),'approved') = 'approved'
                GROUP BY ii.product_id
            ) s ON s.product_id = p.id
            ORDER BY total_profit DESC NULLS LAST, p.name ASC
        """)
        return [row_to_dict(r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{product_id}")
async def get_product(product_id: UUID, user=Depends(get_current_user)):
    pool = await get_pool()

    try:
        row = await pool.fetchrow(
            """
            SELECT p.*, wc.name AS category_name, wc.icon AS category_icon
            FROM products p
            LEFT JOIN warehouse_categories wc ON wc.id = p.category_id
            WHERE p.id=$1
            """,
            product_id,
        )

        if not row:
            raise HTTPException(status_code=404, detail="الصنف غير موجود")

        return row_to_dict(row)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{product_id}")
async def update_product(product_id: UUID, data: ProductRequest, user=Depends(get_current_user)):
    require_warehouse_editor(user)

    name = clean_text(data.name)
    sku = clean_text(data.sku)
    unit = clean_text(data.unit) or "قطعة"
    min_stock = float(data.min_stock or 0)
    category_id = int(data.category_id) if data.category_id else None

    if not name:
        raise HTTPException(status_code=400, detail="اسم الصنف مطلوب")

    if min_stock < 0:
        raise HTTPException(
            status_code=400, detail="الحد الأدنى لا يمكن أن يكون سالباً")

    if data.final_stock is not None and float(data.final_stock) < 0:
        raise HTTPException(
            status_code=400, detail="الكمية الحالية لا يمكن أن تكون سالبة")

    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                current = await conn.fetchrow(
                    """
                    SELECT id, name, current_stock, properties
                    FROM products
                    WHERE id=$1
                    FOR UPDATE
                    """,
                    product_id,
                )

                if not current:
                    raise HTTPException(
                        status_code=404, detail="الصنف غير موجود")

                await ensure_category_exists(conn, category_id)
                await ensure_sku_unique(conn, sku, exclude_product_id=product_id)

                category_name = None

                if category_id:
                    category_name = await conn.fetchval(
                        "SELECT name FROM warehouse_categories WHERE id=$1",
                        category_id,
                    )

                duplicate = await conn.fetchrow(
                    """
                    SELECT id, name
                    FROM products
                    WHERE LOWER(name) = LOWER($1)
                      AND COALESCE(category_id, 0) = COALESCE($2::int, 0)
                      AND id <> $3
                    LIMIT 1
                    """,
                    name,
                    category_id,
                    product_id,
                )

                if duplicate:
                    raise HTTPException(
                        status_code=409,
                        detail=f"الصنف موجود مسبقاً في هذه الفئة: {duplicate['name']}",
                    )

                old_properties = parse_product_properties(
                    current["properties"])
                new_properties = data.properties or {}
                merged_properties = {
                    **old_properties,
                    **new_properties,
                }
                props_json = json.dumps(merged_properties, ensure_ascii=False)

                cost_price_val = round(float(data.cost_price if data.cost_price is not None else (data.base_price or 0)), 3)
                row = await conn.fetchrow(
                    """
                    UPDATE products
                    SET name=$1, sku=$2, category_id=$3, category=$4,
                        unit=$5, min_stock=$6, properties=$7::jsonb,
                        cost_price=$8, base_price=$8
                    WHERE id=$9
                    RETURNING *
                    """,
                    name, sku, category_id, category_name, unit, min_stock,
                    props_json, cost_price_val, product_id,
                )

                if data.final_stock is not None:
                    old_stock = float(current["current_stock"] or 0)
                    final_stock = float(data.final_stock)
                    diff = final_stock - old_stock

                    if diff != 0:
                        row = await conn.fetchrow(
                            """
                            UPDATE products
                            SET current_stock=$1
                            WHERE id=$2
                            RETURNING *
                            """,
                            final_stock,
                            product_id,
                        )

                        await conn.execute(
                            """
                            INSERT INTO stock_movements
                              (product_id, type, quantity, source_type, notes, created_by)
                            VALUES
                              ($1, $2, $3, 'final_stock_edit', $4, $5)
                            """,
                            product_id,
                            "in" if diff > 0 else "out",
                            abs(diff),
                            f"تعديل الكمية من {old_stock} إلى {final_stock}",
                            user.get("id"),
                        )

                        await insert_audit(
                            conn,
                            user,
                            "تعديل كمية صنف",
                            f"{name}: من {old_stock} إلى {final_stock}",
                        )

                await insert_audit(
                    conn,
                    user,
                    "تعديل بيانات صنف",
                    f"{name}",
                )

                return row_to_dict(row)

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{product_id}/adjust")
async def adjust_stock(product_id: UUID, data: AdjustRequest, user=Depends(get_current_user)):
    require_warehouse_editor(user)

    qty = float(data.quantity)

    if qty == 0:
        raise HTTPException(
            status_code=400, detail="الكمية يجب أن تكون رقماً غير صفر")

    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                current = await conn.fetchrow(
                    """
                    SELECT id, name, current_stock
                    FROM products
                    WHERE id=$1
                    FOR UPDATE
                    """,
                    product_id,
                )

                if not current:
                    raise HTTPException(
                        status_code=404, detail="الصنف غير موجود")

                old_stock = float(current["current_stock"] or 0)
                new_stock = old_stock + qty

                if new_stock < 0:
                    raise HTTPException(
                        status_code=400,
                        detail=f"لا يمكن خصم {abs(qty)} — المخزون الحالي {old_stock}",
                    )

                updated = await conn.fetchrow(
                    """
                    UPDATE products
                    SET current_stock=$1
                    WHERE id=$2
                    RETURNING *
                    """,
                    new_stock,
                    product_id,
                )

                await conn.execute(
                    """
                    INSERT INTO stock_movements
                      (product_id, type, quantity, source_type, notes, created_by)
                    VALUES
                      ($1, $2, $3, 'manual', $4, $5)
                    """,
                    product_id,
                    "in" if qty > 0 else "out",
                    abs(qty),
                    data.notes or "تعديل يدوي",
                    user.get("id"),
                )

                await insert_audit(
                    conn,
                    user,
                    "تعديل مخزون",
                    f"{current['name']}: من {old_stock} إلى {new_stock}",
                )

                return row_to_dict(updated)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{product_id}/movements")
async def get_movements(product_id: UUID, user=Depends(get_current_user)):
    pool = await get_pool()

    try:
        rows = await pool.fetch(
            """
            SELECT sm.*, u.full_name AS user_name,
                   p.invoice_number AS purchase_invoice,
                   p.status AS purchase_status,
                   sup.name AS purchase_supplier,
                   CASE
                     WHEN sm.source_type = 'purchase' AND sm.source_id IS NOT NULL AND p.id IS NULL
                     THEN true ELSE false
                   END AS orphaned
            FROM stock_movements sm
            LEFT JOIN users u ON u.id = sm.created_by
            LEFT JOIN purchases p ON sm.source_type = 'purchase' AND sm.source_id = p.id
            LEFT JOIN suppliers sup ON p.supplier_id = sup.id
            WHERE sm.product_id = $1
            ORDER BY sm.created_at DESC
            LIMIT 100
            """,
            product_id,
        )

        return [row_to_dict(r) for r in rows]

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{product_id}/movements/{movement_id}")
async def delete_stock_movement(product_id: UUID, movement_id: int, user=Depends(get_current_user)):
    require_role(user, "admin")
    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                movement = await conn.fetchrow(
                    "SELECT * FROM stock_movements WHERE id=$1 AND product_id=$2",
                    movement_id, product_id,
                )
                if not movement:
                    raise HTTPException(status_code=404, detail="حركة المخزون غير موجودة")

                qty = float(movement["quantity"] or 0)
                if movement["type"] == "in":
                    await conn.execute(
                        "UPDATE products SET current_stock = GREATEST(0, current_stock - $1) WHERE id = $2",
                        qty, product_id,
                    )
                else:
                    await conn.execute(
                        "UPDATE products SET current_stock = current_stock + $1 WHERE id = $2",
                        qty, product_id,
                    )

                await conn.execute("DELETE FROM stock_movements WHERE id=$1", movement_id)

                updated = await conn.fetchrow("SELECT current_stock FROM products WHERE id=$1", product_id)

                await insert_audit(conn, user, "حذف حركة مخزون",
                    f"حذف حركة #{movement_id} — كمية {qty}")

                return {"success": True, "new_stock": float(updated["current_stock"] or 0)}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{product_id}")
async def delete_product(product_id: UUID, user=Depends(get_current_user)):
    require_role(user, "admin")

    pool = await get_pool()

    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                product = await conn.fetchrow(
                    """
                    SELECT id, name
                    FROM products
                    WHERE id = $1
                    FOR UPDATE
                    """,
                    product_id,
                )

                if not product:
                    raise HTTPException(
                        status_code=404,
                        detail="الصنف غير موجود أو تم حذفه مسبقاً",
                    )

                used_in_invoice = await conn.fetchval(
                    "SELECT EXISTS(SELECT 1 FROM invoice_items WHERE product_id = $1)",
                    product_id,
                )

                used_in_warehouse_invoice = await conn.fetchval(
                    "SELECT EXISTS(SELECT 1 FROM warehouse_invoice_items WHERE product_id = $1)",
                    product_id,
                )

                used_in_purchase = await conn.fetchval(
                    "SELECT EXISTS(SELECT 1 FROM purchase_items WHERE product_id = $1)",
                    product_id,
                )

                if used_in_invoice or used_in_warehouse_invoice or used_in_purchase:
                    raise HTTPException(
                        status_code=409,
                        detail="لا يمكن حذف هذا الصنف لأنه مستخدم في فواتير أو مشتريات. حتى نحافظ على تاريخ الفواتير، عدّلي الصنف أو اجعلي كميته صفر بدل الحذف.",
                    )

                await conn.execute(
                    "DELETE FROM stock_movements WHERE product_id = $1",
                    product_id,
                )

                await conn.execute(
                    "DELETE FROM products WHERE id = $1",
                    product_id,
                )

                await insert_audit(
                    conn,
                    user,
                    "حذف صنف مستودع",
                    f"حذف صنف: {product['name']}",
                )

        return {
            "success": True,
            "message": "تم حذف الصنف من قاعدة البيانات بنجاح",
            "deleted_id": str(product_id),
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
